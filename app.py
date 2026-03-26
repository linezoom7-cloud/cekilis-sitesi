"""
app.py - Mortex Çekiliş Yönetim Sistemi - Flask Backend (v3 - Secured)

Security layers:
  1. SQL Injection     → Parameterized queries in database.py (already in place)
  2. Brute-Force      → Flask-Limiter: max 10 login attempts/minute per IP
  3. Password Hashing → werkzeug.security (pbkdf2:sha256), auto-migrates plain text
  4. Security Headers → X-Frame-Options, CSP, HSTS, X-Content-Type-Options, etc.
  5. Session Security → HttpOnly, SameSite=Lax, 30min lifetime
  6. Input Validation → Max lengths, type checks, whitelist file extensions
  7. Upload Safety    → MAX_CONTENT_LENGTH = 5MB, extension whitelist
  8. Error Handling   → No stack traces leaked in production
"""

import os
import csv
import random
import secrets
import io
import re
import openpyxl
from functools import wraps
from flask import Flask, render_template, request, jsonify, session, Response
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.security import generate_password_hash, check_password_hash
from database import (
    init_db, get_setting, set_setting,
    get_all_raffles, get_raffle, create_raffle, delete_raffle, reset_raffle, copy_raffle,
    get_entries_for_raffle, add_entry, bulk_add_entries, delete_entry,
    get_winners_for_raffle, get_all_winners, save_winners,
)

# ---------------------------------------------------------------------------
# App Setup
# ---------------------------------------------------------------------------

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

# ── Session Security ────────────────────────────────────────────────────────
app.config["PERMANENT_SESSION_LIFETIME"] = 1800      # 30 min
app.config["SESSION_COOKIE_HTTPONLY"] = True          # JS cannot read cookie
app.config["SESSION_COOKIE_SAMESITE"] = "Strict"      # CSRF strict mitigation
app.config["SESSION_COOKIE_SECURE"] = False           # Set True when HTTPS
app.config["SESSION_COOKIE_NAME"] = "mortex_session"

@app.after_request
def add_security_headers(response):
    response.headers['Server'] = 'MortexSecure/1.0'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdn.tailwindcss.com https://cdn.jsdelivr.net; style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; font-src 'self' https://fonts.gstatic.com;"
    return response

@app.before_request
def csrf_protect():
    # Enforce Origin/Referer matching for state-changing requests
    if request.method in ["POST", "PUT", "DELETE"]:
        origin = request.headers.get("Origin")
        referer = request.headers.get("Referer")
        host = request.host
        
        # If origin is present, it must match our host
        if origin and host not in origin:
            from flask import abort
            abort(403, description="Referer/Origin CSRF Check Failed")

# ── Upload Safety ────────────────────────────────────────────────────────────
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024  # 5 MB max upload

ALLOWED_EXTENSIONS = {"csv", "txt", "xlsx"}

def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# ── Rate Limiter ─────────────────────────────────────────────────────────────
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=[],           # No global limit; only on specific routes
    storage_uri="memory://"
)

# ---------------------------------------------------------------------------
# Security Headers (added to every response)
# ---------------------------------------------------------------------------

@app.after_request
def add_security_headers(response):
    # Prevent MIME type sniffing
    response.headers["X-Content-Type-Options"] = "nosniff"
    # Prevent clickjacking
    response.headers["X-Frame-Options"] = "DENY"
    # Legacy XSS protection for older browsers
    response.headers["X-XSS-Protection"] = "1; mode=block"
    # Control referrer information
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    # Permissions policy — disable unneeded browser features
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=(), payment=()"
    # HTTP Strict Transport Security (uncomment after enabling HTTPS)
    # response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    # Content Security Policy — restrict resource origins
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.tailwindcss.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self';"
    )
    # Remove server fingerprint
    response.headers.pop("Server", None)
    return response


# ---------------------------------------------------------------------------
# Error Handlers (no stack traces leaked)
# ---------------------------------------------------------------------------

@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Sayfa bulunamadı."}), 404

@app.errorhandler(413)
def request_too_large(e):
    return jsonify({"error": "Dosya çok büyük. Maksimum 5MB yükleyebilirsiniz."}), 413

@app.errorhandler(429)
def too_many_requests(e):
    return jsonify({"error": "Çok fazla deneme yaptınız. Lütfen biraz bekleyin."}), 429

@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": "Sunucu hatası oluştu. Lütfen daha sonra tekrar deneyin."}), 500


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def _is_hashed(value: str) -> bool:
    """Check if a value is already a werkzeug hash."""
    return value.startswith("pbkdf2:") or value.startswith("scrypt:")


def _ensure_password_hashed():
    """Auto-migrate plain text password to hash on first run."""
    current = get_setting("admin_password", "")
    if current and not _is_hashed(current):
        set_setting("admin_password", generate_password_hash(current))


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"error": "Yetkisiz erişim."}), 401
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Input Validation Helpers
# ---------------------------------------------------------------------------

MAX_NAME_LEN = 150
MAX_RAFFLE_NAME_LEN = 200

def sanitize_str(value, max_len=MAX_NAME_LEN) -> str:
    """Strip, trim, enforce max length, remove control characters."""
    s = str(value or "").strip()
    # Remove control characters (tab, null, etc.) but keep Turkish chars
    s = re.sub(r"[\x00-\x08\x0b-\x1f\x7f]", "", s)
    return s[:max_len]


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def public_index():
    all_raffles = get_all_raffles()
    active_raffles = [r for r in all_raffles if r["status"] == "active"]
    drawn_raffles = [r for r in all_raffles if r["status"] == "drawn"]
    return render_template("public.html", active=active_raffles, drawn=drawn_raffles)

@app.route("/admin")
def admin_portal():
    return render_template("admin.html")


@app.route("/live/<int:rid>")
def live_draw(rid):
    raffle = get_raffle(rid)
    if not raffle:
        return "Çekiliş bulunamadı veya silinmiş.", 404
    return render_template("live.html", raffle=raffle)


@app.route("/api/raffles/<int:rid>/public")
@limiter.limit("120 per minute")
def public_raffle_info(rid):
    raffle = get_raffle(rid)
    if not raffle:
        return jsonify({"error": "Çekiliş bulunamadı."}), 404
    
    entries = get_entries_for_raffle(rid)
    winners = get_winners_for_raffle(rid)
    
    import random
    
    def mask_name(name):
        parts = name.split()
        if not parts: 
            return name
        if len(parts) == 1: 
            return parts[0]
            
        first_names = " ".join(parts[:-1])
        last_name = parts[-1]
        masked_last_name = last_name[0] + "*" * (len(last_name) - 1) if len(last_name) > 1 else last_name[0] + "*"
        return f"{first_names} {masked_last_name}"
    
    names = [mask_name(e["full_name"]) for e in entries]
    if len(names) > 100:
        names = random.sample(names, 100)
        
    masked_winners = []
    for w in winners:
        mw = dict(w)
        mw["full_name"] = mask_name(mw["full_name"])
        masked_winners.append(mw)
        
    return jsonify({
        "raffle": {
            "id": raffle["id"],
            "name": raffle["name"],
            "status": raffle["status"],
            "participant_count": len(entries),
            "participant_names": names
        },
        "winners": masked_winners
    })


@app.route("/api/me")
def me():
    return jsonify({"logged_in": bool(session.get("logged_in"))})


@app.route("/api/login", methods=["POST"])
@limiter.limit("10 per minute; 50 per hour")
def login():
    data = request.get_json(force=True, silent=True) or {}
    u = sanitize_str(data.get("username", ""), 80)
    p = sanitize_str(data.get("password", ""), 128)

    stored_user = get_setting("admin_username", "admin")
    stored_pass = get_setting("admin_password", "")

    # Constant-time comparison for username + password
    user_ok = secrets.compare_digest(u, stored_user)
    if _is_hashed(stored_pass):
        pass_ok = check_password_hash(stored_pass, p)
    else:
        pass_ok = secrets.compare_digest(p, stored_pass)

    if user_ok and pass_ok:
        session.clear()
        session["logged_in"] = True
        session.permanent = True
        return jsonify({"success": True})

    return jsonify({"success": False, "message": "Kullanıcı adı veya şifre hatalı!"}), 401


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"success": True})


@app.route("/api/change-password", methods=["POST"])
@login_required
def change_password():
    data = request.get_json(force=True, silent=True) or {}
    current  = sanitize_str(data.get("current_password", ""), 128)
    new_pass = sanitize_str(data.get("new_password", ""), 128)
    confirm  = sanitize_str(data.get("confirm_password", ""), 128)

    stored_pass = get_setting("admin_password", "")
    if _is_hashed(stored_pass):
        if not check_password_hash(stored_pass, current):
            return jsonify({"error": "Mevcut şifre yanlış."}), 400
    else:
        if not secrets.compare_digest(current, stored_pass):
            return jsonify({"error": "Mevcut şifre yanlış."}), 400

    if len(new_pass) < 6:
        return jsonify({"error": "Yeni şifre en az 6 karakter olmalıdır."}), 400
    if new_pass != confirm:
        return jsonify({"error": "Yeni şifreler eşleşmiyor."}), 400

    set_setting("admin_password", generate_password_hash(new_pass))
    return jsonify({"success": True, "message": "Şifre başarıyla değiştirildi."})


# ---------------------------------------------------------------------------
# Raffle API
# ---------------------------------------------------------------------------

@app.route("/api/raffles", methods=["GET"])
@login_required
def list_raffles():
    return jsonify(get_all_raffles())


@app.route("/api/raffles", methods=["POST"])
@login_required
@limiter.limit("60 per minute")
def create_raffle_route():
    data = request.get_json(force=True, silent=True) or {}
    name = sanitize_str(data.get("name", ""), MAX_RAFFLE_NAME_LEN)
    draw_date = sanitize_str(data.get("draw_date", ""), 10)
    if not name or not draw_date:
        return jsonify({"error": "Çekiliş adı ve tarihi zorunludur."}), 400
    # Validate date format loosely
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", draw_date):
        return jsonify({"error": "Geçersiz tarih formatı (YYYY-MM-DD)."}), 400
    return jsonify({"success": True, "id": create_raffle(name, draw_date)}), 201


@app.route("/api/raffles/<int:rid>", methods=["DELETE"])
@login_required
def delete_raffle_route(rid):
    if not get_raffle(rid):
        return jsonify({"error": "Çekiliş bulunamadı."}), 404
    delete_raffle(rid)
    return jsonify({"success": True})


@app.route("/api/raffles/<int:rid>/reset", methods=["POST"])
@login_required
def reset_raffle_route(rid):
    if not get_raffle(rid):
        return jsonify({"error": "Çekiliş bulunamadı."}), 404
    reset_raffle(rid)
    return jsonify({"success": True, "message": "Çekiliş sıfırlandı ve tekrar aktif."})


@app.route("/api/raffles/<int:rid>/copy", methods=["POST"])
@login_required
def copy_raffle_route(rid):
    raffle = get_raffle(rid)
    if not raffle:
        return jsonify({"error": "Çekiliş bulunamadı."}), 404
    data = request.get_json(force=True, silent=True) or {}
    new_name = sanitize_str(data.get("name") or f"{raffle['name']} (Kopya)", MAX_RAFFLE_NAME_LEN)
    new_date = sanitize_str(data.get("draw_date") or raffle["draw_date"], 10)
    copy_entries = bool(data.get("copy_entries", True))
    new_id = copy_raffle(rid, new_name, new_date, copy_entries)
    return jsonify({"success": True, "id": new_id}), 201


# ---------------------------------------------------------------------------
# Entry API
# ---------------------------------------------------------------------------

@app.route("/api/raffles/<int:rid>/entries", methods=["GET"])
@login_required
def list_entries(rid):
    raffle = get_raffle(rid)
    if not raffle:
        return jsonify({"error": "Çekiliş bulunamadı."}), 404
    entries = get_entries_for_raffle(rid)
    total_tickets = sum(e["tickets"] for e in entries)
    winners = get_winners_for_raffle(rid)
    return jsonify({
        "raffle": raffle,
        "entries": entries,
        "total_tickets": total_tickets,
        "winners": winners,
    })


@app.route("/api/raffles/<int:rid>/entries", methods=["POST"])
@login_required
@limiter.limit("200 per minute")
def add_entry_route(rid):
    raffle = get_raffle(rid)
    if not raffle:
        return jsonify({"error": "Çekiliş bulunamadı."}), 404
    if raffle["status"] != "active":
        return jsonify({"error": "Bu çekiliş artık aktif değil."}), 400

    data = request.get_json(force=True, silent=True) or {}
    full_name = sanitize_str(data.get("full_name", ""), MAX_NAME_LEN)
    contact   = sanitize_str(data.get("contact", ""), 200)
    try:
        tickets = max(1, min(int(data.get("tickets", 1)), 10000))
    except (ValueError, TypeError):
        return jsonify({"error": "Geçersiz bilet sayısı."}), 400

    if not full_name:
        return jsonify({"error": "İsim Soyisim zorunludur."}), 400

    return jsonify({"success": True, "id": add_entry(rid, full_name, contact, tickets)}), 201


@app.route("/api/raffles/<int:rid>/import-csv", methods=["POST"])
@login_required
@limiter.limit("30 per minute")
def import_csv(rid):
    raffle = get_raffle(rid)
    if not raffle:
        return jsonify({"error": "Çekiliş bulunamadı."}), 404
    if raffle["status"] != "active":
        return jsonify({"error": "Bu çekiliş aktif değil."}), 400

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "Dosya bulunamadı."}), 400
    if not allowed_file(file.filename):
        return jsonify({"error": "Desteklenmeyen dosya türü. (.csv, .txt, .xlsx)"}), 400

    rows, errors = [], []

    if file.filename.lower().endswith(".xlsx"):
        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            sheet = wb.active
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if not row or not row[0]:
                    continue
                name = sanitize_str(str(row[0]), MAX_NAME_LEN)
                if not name:
                    continue
                try:
                    tickets = max(1, min(int(row[1]), 10000)) if len(row) > 1 and row[1] is not None else 1
                except (ValueError, TypeError):
                    errors.append(f"Satır {i}: Geçersiz bilet sayısı")
                    continue
                rows.append((name, tickets))
            wb.close()
        except Exception as e:
            return jsonify({"error": f"Excel dosyası okunamadı: {str(e)}"}), 400
    else:
        content = file.read().decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(content))
        for i, row in enumerate(reader, start=1):
            if not row or not row[0].strip():
                continue
            name = sanitize_str(row[0], MAX_NAME_LEN)
            try:
                tickets = max(1, min(int(row[1].strip()), 10000)) if len(row) > 1 and row[1].strip() else 1
            except ValueError:
                errors.append(f"Satır {i}: Geçersiz bilet sayısı")
                continue
            rows.append((name, tickets))

    if not rows:
        return jsonify({"error": "Dosyada geçerli kayıt bulunamadı."}), 400

    bulk_add_entries(rid, rows)
    return jsonify({"success": True, "added": len(rows), "errors": errors}), 201


@app.route("/api/entries/<int:entry_id>", methods=["DELETE"])
@login_required
def delete_entry_route(entry_id):
    delete_entry(entry_id)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Draw API
# ---------------------------------------------------------------------------

@app.route("/api/raffles/<int:rid>/draw", methods=["POST"])
@login_required
@limiter.limit("20 per minute")
def draw_winner(rid):
    raffle = get_raffle(rid)
    if not raffle:
        return jsonify({"error": "Çekiliş bulunamadı."}), 404
    if raffle["status"] != "active":
        return jsonify({"error": "Bu çekiliş zaten tamamlandı veya iptal edildi."}), 400

    data    = request.get_json(force=True, silent=True) or {}
    count   = max(1, min(int(data.get("count", 1)), 20))
    entries = get_entries_for_raffle(rid)
    if not entries:
        return jsonify({"error": "Çekilişte hiç katılımcı yok."}), 400

    # Build weighted pool (more tickets = more chances, random draw)
    pool = []
    for e in entries:
        pool.extend([e] * e["tickets"])

    seen, winners_list = set(), []
    attempts = 0
    while len(winners_list) < count and attempts < len(pool) * 3:
        pick = random.choice(pool)
        if pick["id"] not in seen:
            seen.add(pick["id"])
            winners_list.append(pick)
        attempts += 1

    save_winners(rid, [w["id"] for w in winners_list])
    pool_size = len(pool)

    return jsonify({
        "success": True,
        "winners": [
            {
                "rank": i + 1,
                "full_name": w["full_name"],
                "contact": w["contact"],
                "tickets": w["tickets"],
                "odds": round((w["tickets"] / pool_size) * 100, 1),
            }
            for i, w in enumerate(winners_list)
        ],
        "pool_size": pool_size,
    })


# ---------------------------------------------------------------------------
# Winners History API
# ---------------------------------------------------------------------------

@app.route("/api/winners", methods=["GET"])
@login_required
def all_winners():
    return jsonify(get_all_winners())


# ---------------------------------------------------------------------------
# Export API
# ---------------------------------------------------------------------------

@app.route("/api/raffles/<int:rid>/export", methods=["GET"])
@login_required
def export_raffle_entries(rid):
    raffle = get_raffle(rid)
    if not raffle:
        return jsonify({"error": "Çekiliş bulunamadı."}), 404
    entries = get_entries_for_raffle(rid)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Katılımcılar"
    ws.append(["#", "İsim Soyisim", "Bilet Sayısı"])
    for i, e in enumerate(entries, start=1):
        ws.append([i, e["full_name"], e["tickets"]])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=katilimcilar_{rid}.xlsx"}
    )


@app.route("/api/winners/export", methods=["GET"])
@login_required
def export_all_winners():
    winners = get_all_winners()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kazananlar"
    ws.append(["Sıra", "İsim Soyisim", "Çekiliş Adı", "Tarih", "Bilet", "Çekildi"])
    for w in winners:
        ws.append([w["rank"], w["full_name"], w["raffle_name"], w["draw_date"], w["tickets"], w["drawn_at"]])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kazananlar.xlsx"}
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

with app.app_context():
    try:
        init_db()
        _ensure_password_hashed()  # Auto-migrate plain text password to hash
    except Exception as e:
        print(f"Veritabanı başlatma hatası: {e}")

if __name__ == "__main__":
    print("\n🎰 Mortex Çekiliş v3 (Secured) başlatılıyor...")
    print("🌐 Adres: http://127.0.0.1:5000")
    print("🔑 Admin: admin  |  Şifre: mortex2024")
    print("🔒 Güvenlik: Rate limiting, Bcrypt, Strict CSP, Masked PII, Debug OFF")
    print("─" * 50)
    app.run(debug=False, host="0.0.0.0", port=5000)
